"""
Çevirmə məntiqi: mənbə DataFrame-dən şablona məlumat yazır.
Rəqəm sütunları Excel-də rəqəm formatında yazılır.
"""

import pandas as pd
from copy import copy as copy_obj
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO

# Rəqəm formatında yazılacaq sütun açar sözləri
NUMERIC_KEYWORDS = [
    'statistik', 'dəyər', 'value',
    'ədv', 'vat',
    'rüsum', 'duty',
    'gömrük', 'aksiz',
    'miqdar', 'quantity',
    'çəki', 'weight',
    'netto', 'brutto',
    'qiymət', 'məbləğ', 'cəmi',
]


def _is_numeric_col(col_name: str) -> bool:
    col_lower = str(col_name).lower()
    return any(kw in col_lower for kw in NUMERIC_KEYWORDS)


def _to_number(val):
    """Dəyəri rəqəmə çevir — mümkün deyilsə orijinalı qaytar."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return None if pd.isna(val) else float(val)
    try:
        cleaned = str(val).strip().replace(' ', '').replace(',', '.')
        return float(cleaned)
    except Exception:
        return val


def fill_template(source_df: pd.DataFrame,
                  template_bytes: bytes,
                  mapping: dict,
                  header_row: int = 1) -> bytes:
    """Şablonu açır, məlumatları yazır, formatı qoruyur."""
    wb = load_workbook(BytesIO(template_bytes))
    ws = wb.active

    template_headers = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            template_headers[str(cell.value).strip()] = cell.column

    last_row = ws.max_row
    if last_row > header_row:
        ws.delete_rows(header_row + 1, last_row - header_row)

    style_refs = {col: ws.cell(row=header_row, column=idx)
                  for col, idx in template_headers.items()}

    for row_idx, (_, src_row) in enumerate(source_df.iterrows(), start=header_row + 1):
        for tmpl_col, src_col in mapping.items():
            col_idx = template_headers.get(tmpl_col)
            if col_idx is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)

            if src_col == "(boş bur)":
                cell.value = ""
            else:
                val = src_row.get(src_col, "")
                if isinstance(val, float) and pd.isna(val):
                    val = None
                if _is_numeric_col(tmpl_col) or _is_numeric_col(src_col):
                    num = _to_number(val)
                    if isinstance(num, float):
                        cell.value         = num
                        cell.number_format = '#,##0.00'
                    else:
                        cell.value = val
                else:
                    cell.value = val

            ref = style_refs.get(tmpl_col)
            if ref:
                if ref.font:      cell.font      = copy_obj(ref.font)
                if ref.fill:      cell.fill      = copy_obj(ref.fill)
                if ref.border:    cell.border    = copy_obj(ref.border)
                if ref.alignment: cell.alignment = copy_obj(ref.alignment)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    """
    DataFrame-i Excel faylına çevirir.
    Rəqəm sütunları avtomatik #,##0.00 formatında yazılır.
    """
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
        wb  = writer.book
        # Sheet adı nə olursa olsun götür
        ws  = wb.worksheets[0]

        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)

            # Sütun enini məzmuna görə avtomatik tənzimlə
            max_len = max(
                (len(str(v)) for v in df[col_name].dropna()),
                default=len(str(col_name))
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

            if not _is_numeric_col(col_name):
                continue

            for row_idx in range(2, len(df) + 2):
                cell = ws[f"{col_letter}{row_idx}"]
                num  = _to_number(cell.value)
                if isinstance(num, float):
                    cell.value         = num
                    cell.number_format = '#,##0.00'

    return buf.getvalue()
