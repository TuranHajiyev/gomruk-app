"""
Gömrük Ödənişlərinin Uyğunlaşdırılması Modulu.

Mədaxil sütunundakı ödənişləri Məxaric sütunundakı
bəyannamə silmələri ilə uyğunlaşdırır.

Qayda:
  - "Ödənilib" yazılıbsa → yalnız o mədaxili uyğunlaşdır
  - Boşdursa → bütün mədaxilləri uyğunlaşdır
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import re


# ── Fayl oxuma ────────────────────────────────────────────

def _to_float(val):
    """'1 215.00' → 1215.0"""
    if val is None:
        return 0.0
    s = str(val).replace(' ', '').replace(',', '.').strip()
    try:
        return float(s)
    except Exception:
        return 0.0


def parse_file(file_bytes: bytes) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """
    Faylı oxuyur.
    Qaytarır:
        madaxil_df  — Mədaxil sətirləri
        mexaric_df  — Məxaric sətirləri
        meta        — {shirket, dovr}
    """
    wb  = load_workbook(BytesIO(file_bytes), data_only=True)
    ws  = wb.active
    rows = list(ws.iter_rows(values_only=True))

    meta = {'shirket': '', 'dovr': ''}
    madaxil_rows = []
    mexaric_rows = []

    mode = None   # 'madaxil' | 'mexaric'

    for i, row in enumerate(rows):
        cells = {j: (str(v).strip() if v is not None else '') for j, v in enumerate(row)}
        line  = ' '.join(cells.values())

        # Meta: şirkət adı
        for v in cells.values():
            if 'Məhdud' in v or 'Cəmiyyəti' in v or 'MMC' in v or 'ASC' in v:
                meta['shirket'] = v.split('\n')[0].strip()
            if 'dən' in v and 'dək' in v and re.search(r'\d{2}\.\d{2}\.\d{4}', v):
                meta['dovr'] = v.strip()

        # Bölmə başlıqlarını tap
        if 'Mədaxil' in line and 'Məxaric' not in line:
            mode = 'madaxil'
            continue
        if 'Məxaric' in line:
            mode = 'mexaric'
            continue

        # Footer sətirini keç
        if '©' in line or 'Dövrün' in line or 'Hesabat dövrü' in line:
            continue

        # Başlıq sətirini keç
        if 'Əməliyyat tarixi' in line or 'Hesab Tarixi' in line:
            continue

        # ── Mədaxil sətirləri ─────────────────────────
        if mode == 'madaxil':
            # Sütunlar: col3=tarix, col11=qəbz, col14=təyinat, col18=növ, col22=məbləğ, col32=ödənilib
            tarix    = cells.get(3, '') or cells.get(4, '')
            qebz     = cells.get(11, '') or cells.get(12, '')
            teyinat  = cells.get(14, '') or cells.get(15, '')
            nov      = cells.get(18, '') or cells.get(19, '')
            meblег   = cells.get(22, '') or cells.get(23, '')
            odenilib = cells.get(32, '') or cells.get(33, '')

            if tarix and teyinat and meblег:
                madaxil_rows.append({
                    'Tarix':            tarix,
                    'Qəbz Nömrəsi':     qebz,
                    'Ödənişin Təyinatı': teyinat,
                    'Ödənişin Növü':     nov,
                    'Məbləğ':           _to_float(meblег),
                    'Status':           odenilib.strip(),
                })

        # ── Məxaric sətirləri ─────────────────────────
        elif mode == 'mexaric':
            # Sütunlar: col1=tarix, col9=hesab, col13=əməliyyat növü, col17=məbləğ, col20=bəyannamə, col25=prosedur
            tarix     = cells.get(1, '') or cells.get(2, '')
            hesab     = cells.get(9, '') or cells.get(10, '')
            emel_nov  = cells.get(13, '') or cells.get(14, '')
            meblег    = cells.get(17, '') or cells.get(18, '')
            beyanname = cells.get(20, '') or cells.get(21, '')
            prosedur  = cells.get(25, '') or cells.get(26, '')

            if tarix and emel_nov and meblег and beyanname:
                mexaric_rows.append({
                    'Tarix':           tarix,
                    'Hesab':           hesab,
                    'Əməliyyat Növü':  emel_nov,
                    'Məbləğ':          _to_float(meblег),
                    'Bəyannamə':       beyanname.strip(),
                    'Prosedur':        prosedur,
                })

    madaxil_df = pd.DataFrame(madaxil_rows)
    mexaric_df = pd.DataFrame(mexaric_rows)
    return madaxil_df, mexaric_df, meta


# ── Uyğunlaşdırma məntiqi ────────────────────────────────

def uygunlasdir(madaxil_df: pd.DataFrame,
                mexaric_df: pd.DataFrame,
                yalniz_odenilib: bool) -> pd.DataFrame:
    """
    Hər mədaxil ödənişini müvafiq məxaric sətirləri ilə uyğunlaşdırır.

    Uyğunlaşdırma qaydası:
      1. Ödənişin Təyinatı = Məxaric Hesabı (ƏDV depozit, Avans, Xidmətlər)
      2. Məbləğlər cəmi uyğun gəlir
    """
    if madaxil_df.empty or mexaric_df.empty:
        return pd.DataFrame()

    # Yalnız "Ödənilib" filteri
    if yalniz_odenilib:
        src = madaxil_df[madaxil_df['Status'].str.strip() == 'Ödənilib'].copy()
    else:
        src = madaxil_df.copy()

    if src.empty:
        return pd.DataFrame()

    # Hesab → Ödənişin Təyinatı uyğunluq xəritəsi
    hesab_map = {
        'ƏDV depozit':          ['ƏDV depozit', 'ƏDV'],
        'Avans':                ['Avans'],
        'Xidmətlər':            ['Xidmətlər', 'Xidmətlər üzrə ƏDV'],
        'Xidmətlər üzrə ƏDV':  ['Xidmətlər', 'Xidmətlər üzrə ƏDV'],
        'ƏDV':                  ['ƏDV depozit', 'ƏDV'],
    }

    results = []

    for _, mad_row in src.iterrows():
        teyinat = mad_row['Ödənişin Təyinatı']
        meblег  = mad_row['Məbləğ']

        # Bu mədaxilə uyğun məxaric sətirləri tap
        # Hesab uyğunluğu
        uygun_hesablar = hesab_map.get(teyinat, [teyinat])
        mexaric_sub = mexaric_df[
            mexaric_df['Hesab'].isin(uygun_hesablar)
        ].copy()

        if mexaric_sub.empty:
            # Hesab uyğunluğu olmasa — ödənişin təyinatı ilə əməliyyat növünü uyğunlaşdır
            mexaric_sub = mexaric_df[
                mexaric_df['Əməliyyat Növü'].str.contains(teyinat, case=False, na=False)
            ].copy()

        # Məbləğ cəmi uyğunluğu yoxla
        mexaric_sum = mexaric_sub['Məbləğ'].sum()
        mebleg_uygun = abs(mexaric_sum - meblег) < 0.02

        # Hər bəyannamə üçün cəm
        if not mexaric_sub.empty:
            beyanname_cemler = (
                mexaric_sub.groupby('Bəyannamə')['Məbləğ']
                .sum()
                .reset_index()
                .rename(columns={'Məbləğ': 'Bəyannamə Üzrə Məbləğ'})
            )
            beyanname_list = ', '.join(mexaric_sub['Bəyannamə'].unique())
            beyanname_sayi = mexaric_sub['Bəyannamə'].nunique()
        else:
            beyanname_list  = '—'
            beyanname_sayi  = 0
            mexaric_sum     = 0.0
            mebleg_uygun    = False

        results.append({
            'Tarix':              mad_row['Tarix'],
            'Qəbz Nömrəsi':      mad_row['Qəbz Nömrəsi'],
            'Ödənişin Təyinatı':  teyinat,
            'Ödənişin Növü':      mad_row['Ödənişin Növü'],
            'Mədaxil Məbləği':    meblег,
            'Məxaric Cəmi':       round(mexaric_sum, 2),
            'Fərq':               round(meblег - mexaric_sum, 2),
            'Uyğunluq':           '✅ Uyğun' if mebleg_uygun else '⚠️ Fərq var',
            'Bəyannamə Sayı':     beyanname_sayi,
            'Bəyannamələr':       beyanname_list,
            'Status':             mad_row['Status'] or '—',
        })

    return pd.DataFrame(results)


def to_excel_result(
        madaxil_df: pd.DataFrame,
        mexaric_df: pd.DataFrame,
        uygun_df: pd.DataFrame) -> bytes:
    """Nəticəni 3 vərəqli Excel faylına yaz."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    thin  = Side(style='thin', color='CCCCCC')
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, cols):
        hdr_fill = PatternFill('solid', fgColor='1F4E79')
        hdr_font = Font(bold=True, color='FFFFFF', size=10)
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font      = hdr_fill and hdr_font
            c.fill      = hdr_fill
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border    = brd
        ws.row_dimensions[1].height = 28

    def write_df(ws, df):
        style_header(ws, list(df.columns))
        alt = PatternFill('solid', fgColor='EBF3FB')
        for ri, (_, row) in enumerate(df.iterrows(), 2):
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border    = brd
                c.alignment = Alignment(vertical='center')
                if ri % 2 == 0:
                    c.fill = alt
                if isinstance(val, float):
                    c.number_format = '#,##0.00'
        for ci, col in enumerate(df.columns, 1):
            max_w = max(
                (len(str(v)) for v in df[col].dropna()),
                default=len(col))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 4, 45)

    # Vərəq 1: Uyğunlaşdırma nəticəsi
    ws1 = wb.active
    ws1.title = 'Uyğunlaşdırma'
    write_df(ws1, uygun_df)

    # Vərəq 2: Mədaxil
    ws2 = wb.create_sheet('Mədaxil')
    write_df(ws2, madaxil_df)

    # Vərəq 3: Məxaric + bəyannamə
    ws3 = wb.create_sheet('Məxaric')
    write_df(ws3, mexaric_df)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Streamlit səhifəsi ────────────────────────────────────

def show():
    st.header("💳 Gömrük Ödənişlərinin Uyğunlaşdırılması")
    st.caption(
        "Mədaxil ödənişlərini məxaric bəyannamə silmələri ilə avtomatik uyğunlaşdırır."
    )

    uploaded = st.file_uploader(
        "Avans hesabı çıxarışını yükləyin (.xlsx)",
        type=["xlsx"],
        key="odenis_uploader"
    )

    if not uploaded:
        st.info("Gömrük Komitəsinin avans hesabı çıxarışını yükləyin.")
        return

    file_bytes = uploaded.read()

    with st.spinner("Fayl oxunur..."):
        try:
            madaxil_df, mexaric_df, meta = parse_file(file_bytes)
        except Exception as e:
            st.error(f"Fayl oxunarkən xəta: {e}")
            return

    # Meta
    if meta.get('shirket'):
        st.info(f"🏢 **{meta['shirket']}**  |  📅 {meta.get('dovr','')}")

    # Statistika
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Mədaxil sətiri",  len(madaxil_df))
    c2.metric("Məxaric sətiri",  len(mexaric_df))
    c3.metric("Mədaxil cəmi",    f"{madaxil_df['Məbləğ'].sum():,.2f} AZN")
    c4.metric("Məxaric cəmi",    f"{mexaric_df['Məbləğ'].sum():,.2f} AZN")

    st.divider()

    # Filtr seçimi
    col_a, col_b = st.columns(2)
    with col_a:
        yalniz_odenilib = st.toggle(
            "Yalnız **Ödənilib** işarəli mədaxilləri uyğunlaşdır",
            value=True,
            help="Aktiv: yalnız 'Ödənilib' sütunu dolub olanlar. Deaktiv: hamısı."
        )
    with col_b:
        teyinat_filter = st.multiselect(
            "Ödənişin Təyinatına görə süzgəc:",
            options=sorted(madaxil_df['Ödənişin Təyinatı'].unique().tolist()),
            default=[],
            placeholder="Hamısı (boş buraxın)"
        )

    # Uyğunlaşdır
    df_src = madaxil_df.copy()
    if teyinat_filter:
        df_src = df_src[df_src['Ödənişin Təyinatı'].isin(teyinat_filter)]

    with st.spinner("Uyğunlaşdırılır..."):
        uygun_df = uygunlasdir(df_src, mexaric_df, yalniz_odenilib)

    if uygun_df.empty:
        st.warning("Uyğunlaşdırılacaq məlumat tapılmadı.")
        return

    # Nəticə statistikası
    st.subheader("📊 Uyğunlaşdırma Nəticəsi")
    r1, r2, r3 = st.columns(3)
    uygun_say   = (uygun_df['Uyğunluq'] == '✅ Uyğun').sum()
    ferq_say    = (uygun_df['Uyğunluq'] == '⚠️ Fərq var').sum()
    r1.metric("Uyğunlaşdırılan",  len(uygun_df))
    r2.metric("✅ Uyğun",          uygun_say)
    r3.metric("⚠️ Fərq var",       ferq_say)

    # Rəng ilə cədvəl
    def highlight(row):
        if row['Uyğunluq'] == '✅ Uyğun':
            return ['background-color: #1a4731; color: #6ee7b7'] * len(row)
        elif row['Uyğunluq'] == '⚠️ Fərq var':
            return ['background-color: #4a1c1c; color: #fca5a5'] * len(row)
        return [''] * len(row)

    st.dataframe(
        uygun_df.style.apply(highlight, axis=1),
        use_container_width=True,
        height=420
    )

    # Bəyannamə üzrə xülasə
    st.divider()
    st.subheader("📋 Bəyannamə Üzrə Xülasə")
    beyanname_xulase = (
        mexaric_df.groupby(['Bəyannamə', 'Əməliyyat Növü'])['Məbləğ']
        .sum()
        .reset_index()
        .rename(columns={'Məbləğ': 'Cəmi Məbləğ'})
        .sort_values('Bəyannamə')
    )
    beyanname_xulase['Cəmi Məbləğ'] = beyanname_xulase['Cəmi Məbləğ'].round(2)
    st.dataframe(beyanname_xulase, use_container_width=True, height=300)

    # Yükləmə
    st.divider()
    dl1, dl2 = st.columns(2)
    with dl1:
        excel_bytes = to_excel_result(madaxil_df, mexaric_df, uygun_df)
        st.download_button(
            "📥 Nəticəni Excel kimi yüklə (3 vərəq)",
            data=excel_bytes,
            file_name="odenis_uygunlasma.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
    with dl2:
        csv = uygun_df.to_csv(index=False).encode('utf-8-sig')
        st.download_button(
            "📄 CSV kimi yüklə",
            data=csv,
            file_name="odenis_uygunlasma.csv",
            mime="text/csv",
            use_container_width=True
        )
